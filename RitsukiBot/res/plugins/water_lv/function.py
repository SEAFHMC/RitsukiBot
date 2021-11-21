import ujson
import requests
import openpyxl


async def get_dc_lv(ShuiWei):
    url = "https://tianqiapi.com/api?version=v6&appid=93561677&appsecret=xTW8n2Al&cityid=101200201"
    r = requests.get(url)
    ujsons = ujson.loads(r.text)
    wb = openpyxl.load_workbook("res/docs/dc_lv_log.xlsx")
    sh0 = wb.worksheets[0]
    sh0.cell(sh0.max_row+1, 1).value = ujsons["update_time"]
    sh0.cell(sh0.max_row, 2).value = ShuiWei
    yesterday_time = sh0.cell(sh0.max_row-1, 1).value[11:]
    yesterday_ShuiWei = round(float(sh0.cell(sh0.max_row-1, 2).value), 2)
    minus = round(round(float(ShuiWei), 2)-yesterday_ShuiWei, 2)
    if float(minus) > 0:
        flag = "水位上涨"+str(int(abs(minus*100)))+"cm\n"
    elif float(minus) < 0:
        flag = "水位下降"+str(int(abs(minus*100)))+"cm\n"
    else:
        flag = "无变化\n"
    wb.save("res/docs/dc_lv_log.xlsx")
    result = "测量时间："+ujsons["update_time"]+"\n"+"测量部位：东汊\n目前水位："+ShuiWei+"m\n"+"较昨日"+yesterday_time+" 测量 "+str(yesterday_ShuiWei)+"m\n"+flag+"当前天气："+ujsons["wea"]+"\n"+"风力："+ujsons["win"]+ujsons["win_speed"]
    wb.close()
    return result


async def get_xc_lv(ShuiWei):
    url = "https://tianqiapi.com/api?version=v6&appid=93561677&appsecret=xTW8n2Al&cityid=101200201"
    r = requests.get(url)
    ujsons = ujson.loads(r.text)
    wb = openpyxl.load_workbook("res/docs/xc_lv_log.xlsx")
    sh0 = wb.worksheets[0]
    sh0.cell(sh0.max_row+1, 1).value = ujsons["update_time"]
    sh0.cell(sh0.max_row, 2).value = ShuiWei
    yesterday_time = sh0.cell(sh0.max_row-1, 1).value[11:]
    yesterday_ShuiWei = round(float(sh0.cell(sh0.max_row-1, 2).value), 2)
    minus = round(round(float(ShuiWei), 2)-yesterday_ShuiWei, 2)
    if float(minus) > 0:
        flag = "水位上涨"+str(int(abs(minus*100)))+"cm\n"
    elif float(minus) < 0:
        flag = "水位下降"+str(int(abs(minus*100)))+"cm\n"
    else:
        flag = "无变化\n"
    wb.save("res/docs/xc_lv_log.xlsx")
    result = "测量时间："+ujsons["update_time"]+"\n"+"测量部位：西汊\n目前水位："+ShuiWei+"m\n"+"较昨日"+yesterday_time+" 测量 "+str(yesterday_ShuiWei)+"m\n"+flag+"当前天气："+ujsons["wea"]+"\n"+"风力："+ujsons["win"]+ujsons["win_speed"]
    wb.close()
    return result
